import math
import os
from datetime import datetime
import openpyxl
import pandas as pd
import helper
import enum

REPORT_HEADER_COLOR = 'FFFFD8B1'    # in HEX8 format

class Report:
    Name = 'Timesheet_Report.xlsx'
    class Header:
        Color = 'FFFFD8B1'    # in HEX8 format
        class Columns(enum.Enum):
            Name = 'Name'
            Day = 'Day'
            HanaHours = 'S4Hana Hours'
            ClientHours = 'Client Hours'
            ClientLeave = 'Leave'       # its value will be based on yellow color in client timesheet

class CTS:
    Folder = 'Client'     # contains all timesheets of different squads
    ActualStartRow = 12  # row starts from 1. It is count of rows not index
    RowProcessStopColumn = 'Total Billable Hours'
    ProcessStop = 'India'
    NamePrefixToRemoveLength = 18
    class ColorCode(enum.Enum):
        Leave = '#FFFF00'
    
    class Columns(enum.Enum):
        Name = 0
        Country = 3
        FirstDate = 4

class STS:
    DumpFile = 'hana.xlsx'
    ActualStartRow = 2 # row starts from 1. It is count of rows not index
    class Columns(enum.Enum):
        Name = 1
        FirstDate = 4

class ClientLeave(enum.Enum):
    Yes = 'Yes'
    No = 'No'

DEFAULT_SHEET_NAME_OF_NEW_WORKBOOK = 'Sheet'
OPENPYXL_ENGINE = 'openpyxl'    # Note: Don't change its value. It is a standard value.
EXCEL_FILE_NAME_LENGTH_LIMIT = 31
EXTRA_CELL_WIDTH = 4

def parseClientTimesheets(directory, hanaResult):
    print('INFO: Start parse timesheets from "%s" folder' % directory)
    wb = openpyxl.Workbook()

    fileCount = helper.getFileCount(directory)
    print('%s files found' % fileCount)

    failedFileCount = 0
    skippedFileCount = 0

    for fileName in os.listdir(directory):
        if not helper.isFileHasSpecifiedExtension(fileName, ['.xlsx', '.xls']):
            skippedFileCount += 1
            print('WARN: "%s" is not an excel file' % fileName)
            continue

        try:
            filePath = os.path.join(directory, fileName)
            dataFrame = pd.read_excel(filePath, skiprows = CTS.ActualStartRow - 1, engine = OPENPYXL_ENGINE) # columns of interest starts from 12th row

            clientDayHoursByName = parseSingleClientTimesheet(dataFrame)
            print('INFO: "%s" parsed successfully' % fileName)

            leavesByName  = getLeavesByName(filePath)
            comparisonResult = compareHanaWithClientDetails(hanaResult, clientDayHoursByName, leavesByName)
            
            filenameWithoutExtension = helper.getFileNameWithoutExtension(fileName)
            sheetName = filenameWithoutExtension[CTS.NamePrefixToRemoveLength:]
            sheetName = sheetName[0:EXCEL_FILE_NAME_LENGTH_LIMIT]

            wb = generateReport(wb, comparisonResult, sheetName)
            print('INFO: Report for "%s" generated successfully' % fileName)
        except Exception as e:
            failedFileCount += 1
            print('ERROR: "%s" occured when processing %s' %(e, fileName))

    reportName = '_'.join([helper.getCurrentMonthName(), Report.Name])
    wb.save(reportName)
    print('INFO: "%s" generated successfully' % reportName)
    print('INFO: Report for Client timesheets: %s SUCCESS, %s FAILED %s SKIPPED' % (fileCount - failedFileCount - skippedFileCount, failedFileCount, skippedFileCount))

def parseHana(dataFrame):
    columns = dataFrame.columns
    dayHoursByName = {}

    for r in range(len(dataFrame.index)):
        name = dataFrame.loc[r, columns[STS.Columns.Name.value]]

        for c in range(STS.Columns.FirstDate.value, len(columns)):
            currentDateColumn = columns[c]

            dayHour = {}
            day = int(currentDateColumn[:2])   # first 2 chars represent day. Column Format : dd.mm.yyyy
            hours = dataFrame.loc[r, currentDateColumn]
            dayHour[day] = 0 if math.isnan(hours) else hours
            dayHoursByName.setdefault(name, []).append(dayHour)
    return dayHoursByName
    
def parseSingleClientTimesheet(dataFrame):
    columns = dataFrame.columns
    totalRows = len(dataFrame.index)
    dayHoursByName = {}

    for r in range(totalRows):
        name = dataFrame.loc[r, columns[CTS.Columns.Name.value]]
        country = dataFrame.loc[r, columns[CTS.Columns.Country.value]]
        
        if(country != CTS.ProcessStop):     # if its a valid row, can also check person number...
            break

        for c in range(CTS.Columns.FirstDate.value, len(columns)):
            if (columns[c] == CTS.RowProcessStopColumn):
                break
            else:                
                fridayOfCurrentWeek = helper.getSpecifDayOfCurrentWeek(4)   # 4 - index of Friday
                today = datetime.today()
                lastDayOfCurrentMonth = helper.getLastDayOfMonth(today.year, today.month)
                checkDay = lastDayOfCurrentMonth if today.day > fridayOfCurrentWeek else fridayOfCurrentWeek

                currentDateColumn = columns[c]
                day = int(currentDateColumn)
                if day > checkDay:    # process data only till checkDay
                    break
                
                dayHour = {}
                hours = dataFrame.loc[r,currentDateColumn]
                dayHour[day] = 0 if math.isnan(hours) else hours
                dayHoursByName.setdefault(name, []).append(dayHour)
    return dayHoursByName

def compareHanaWithClientDetails(hanaResult, clientResult, leavesByName):
    comparisonResult = []
    for person in hanaResult:
        clientResultForPerson = clientResult.get(person)
        if clientResultForPerson is None:
            continue
        for dh in hanaResult[person]:
            [(day, hour)] = dh.items()      # single the object has only one key-value pair
            for dh2 in clientResultForPerson:
                [(day2, hour2)] = dh2.items()
                if day == day2:
                    if (hour != hour2) or (hour == 0 and hour2 == 0):       # either mismatch hours, or hours are empty in both timesheets
                        hasPersonAnyLeaveInTheMonth = leavesByName.get(person)     # First check if name exists in the leave result
                        isPersonOnLeave = ClientLeave.Yes.value if hasPersonAnyLeaveInTheMonth and int(day) in leavesByName[person] else ClientLeave.No.value

                        verifyData = {}
                        reportColumns = Report.Header.Columns
                        verifyData[reportColumns.Name.value] = person
                        verifyData[reportColumns.Day.value] = day
                        verifyData[reportColumns.HanaHours.value] = hour
                        verifyData[reportColumns.ClientHours.value] = hour2
                        verifyData[reportColumns.ClientLeave.value] = isPersonOnLeave
                        comparisonResult.append(verifyData)
                    break
    return comparisonResult

def getLeavesByName(clientTimesheetPath):
    wb = openpyxl.load_workbook(clientTimesheetPath, data_only = True)
    sheet = wb.active
    stop = False
    headerRow = ''
    
    leavesByName = {}
    for row, row_cells in enumerate(sheet.iter_rows()):
        cellNum = 0   
                                 
        if row < CTS.ActualStartRow - 1:    # skip unnecessary rows. -1 becoz row index starts from 0                                                                          
            continue
        elif row == CTS.ActualStartRow - 1: # contains column names for data of interest. -1 becoz row index starts from 0
            headerRow = row_cells
            continue

        for cell in row_cells:
            if cellNum == 0 and cell.value is None: # if the row is not resource row, it means all resource's timings have been processed
                stop = True
                break

            color = helper.getCellColor(cell)
            if color == CTS.ColorCode.Leave.value:
                name = row_cells[CTS.Columns.Name.value].value
                day = headerRow[cellNum].value
                leavesByName.setdefault(name, []).append(day)
                # print(cell.value, color, row_cells[0].value, headerRow[cellNum].value)
            cellNum += 1

        if stop:
            break
    return leavesByName            

def generateReport(wb, reportResult, sheetName) :
    helper.removeDefaultSheet(wb)
    sheet = helper.createSheet(wb, sheetName)
    headerColumns = [c.value for c in Report.Header.Columns]

    helper.renderDataInSheet(sheet, reportResult, headerColumns, Report.Header.Color)
    helper.adjustCellWidthToContent(sheet, EXTRA_CELL_WIDTH) # add EXTRA_CELL_WIDTH becoz otherwise the width does not fit the content correctly
    helper.hideGridLines(sheet)
    helper.applyFilter(sheet)
    return wb

if __name__ == '__main__':
    try:
        print('INFO: "%s" read start' % STS.DumpFile )
        dataFrame = pd.read_excel(STS.DumpFile, skiprows = STS.ActualStartRow - 1, engine = OPENPYXL_ENGINE)   # columns of interest starts from 2nd row
        print('INFO: "%s" read successfully' % STS.DumpFile )

        print('INFO: "%s" parse start' % STS.DumpFile )
        hanaResult = parseHana(dataFrame)
        print('INFO: "%s" parsed successfully' % STS.DumpFile )

        parseClientTimesheets(CTS.Folder, hanaResult)

    except Exception as e:
        print('ERROR: %s' % e)
    