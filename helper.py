import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import calendar
import os
from openpyxl.styles import Alignment

def adjustCellWidthToContent(ws, extraWidth):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + extraWidth

def isFileHasSpecifiedExtension(fileName, extensions):
    return fileName.endswith(tuple(extensions))

def getSpecifDayOfCurrentWeek(dayIndex):
    d = datetime.today()
    while d.weekday() != dayIndex: # date.weekday method represents Monday through Sunday as 0 through 6
        d += timedelta(1)   # increment by a day
    return d.day

def getLastDayOfMonth(year, month):
    return calendar.monthrange(year, month)[1]

def centerAlignCellData(cell):
    cell.alignment = Alignment(horizontal='center')

def getCurrentMonthName():
    return datetime.now().strftime("%B")

def hideGridLines(ws):
    ws.sheet_view.showGridLines = False

def getFileCount(dir):
    return len([name for name in os.listdir(dir)])

def getFileNameWithoutExtension(fileName):
    return os.path.splitext(fileName)[0]

def getCellColor(cell):
    try:
        color = '#' + cell.fill.start_color.rgb[2:]
    except:
        colorIndex = cell.fill.start_color.index
        Colors = openpyxl.styles.colors.COLOR_INDEX
        color = str(Colors[colorIndex])
        color = "#" + color[2:]
    return color

def getFillColor(color):
    return PatternFill(start_color = color, end_color = color, fill_type = 'solid')    

def createHeader(sheet, headerColumns, headerColor):
    headerRowNum = 1
    for i in range(1, len(headerColumns) + 1):
        sheet.cell(headerRowNum, i).value = headerColumns[i-1]
        sheet.cell(headerRowNum, i).fill = getFillColor(headerColor) 

def renderDataInSheet(sheet, data, headerColumns, headerColor):
    createHeader(sheet, headerColumns, headerColor)
    for row in range(len(data)):
        for col in range(1, len(headerColumns) + 1):
            cell = sheet.cell(row+2, col)
            cell.value = data[row][headerColumns[col-1]]
            centerAlignCellData(cell)

def removeDefaultSheet(wb):
    defaultSheet = "Sheet"
    if defaultSheet in wb.sheetnames:
        wb.remove(wb[defaultSheet])

def createSheet(wb, sheetName):
    wb.create_sheet(sheetName)
    return wb[sheetName]

def applyFilter(ws):
    ws.auto_filter.ref = ws.dimensions
